# app/infrastructure/file_converter/file_converter.py
import os
import csv
import io
import tempfile
import asyncio
from pathlib import Path
from typing import Tuple, Callable
from fastapi import UploadFile
import pandas as pd
import openpyxl
# from openpyxl.utils import coordinate_from_string, column_index_from_string
# from openpyxl.cell import Cell
from app.core.config import settings
from app.infrastructure.file_converter.llamaparse_converter import LlamaParseConverter

from app.core.logger import get_logger

logger = get_logger()


class FileConverter:
    """
    Converts files to OpenAI-compatible formats.
    Currently supports:
        - CSV → TXT (comma-separated lines)
        - XLSX/XLS → Markdown (per sheet)
    """

    def __init__(self):
        self._handlers: dict[str, Callable] = {
            ".csv": self._convert_csv_to_txt,
            ".xlsx": self._convert_excel_to_md,
            ".xls": self._convert_excel_to_md,
            ".xlsm": self._convert_excel_to_md,
        }

        self._llama: LlamaParseConverter | None = None
        api = getattr(settings, "LLAMACLOUD_API_KEY", None)
        enable = getattr(settings, "LLAMAPARSE_ENABLE", True)
        api_val = api.get_secret_value() if hasattr(api, "get_secret_value") else api

        if enable and api_val:
            try:
                self._llama = LlamaParseConverter(api_key=api_val, verbose=False)
            except Exception as e:
                logger.warning("Failed to initialize LlamaParse converter, will fallback to local: %s", e)
        else:
            logger.info("LlamaParse disabled or API key missing; using local Excel converter")

    async def convert(self, source_path: Path, filename: str) -> Tuple[Path, str]:
        """
        Convert a file to OpenAI-compatible format if needed.

        Args:
            source_path: Path to original file
            filename: Original filename (to detect extension)

        Returns:
            (converted_path, new_filename) or (source_path, filename) if no conversion needed
        """
        ext = os.path.splitext(filename.lower())[-1]
        handler = self._handlers.get(ext)

        if not handler:
            logger.debug("No conversion needed for %s", ext)
            return source_path, filename

        logger.info("Converting %s (%s) to OpenAI-compatible format", filename, ext)
        try:
            # Run blocking conversion in thread
            result = await asyncio.to_thread(handler, source_path, filename)
            return result
        except Exception as e:
            logger.exception("Conversion failed for %s: %s", filename, e)
            raise RuntimeError(f"File conversion failed: {e}") from e

    # ---------- CSV -> TXT ----------
    def _convert_csv_to_txt(self, source_path: Path, filename: str) -> Tuple[Path, str]:
        with open(source_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            txt_content = "\n".join([", ".join(row) for row in reader])

        new_filename = f"{os.path.splitext(filename)[0]}.txt"
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".txt")
        tmp_path = Path(tmp_file.name)

        with tmp_file:
            tmp_file.write(txt_content.encode("utf-8"))

        logger.debug("CSV converted to TXT: %s -> %s", filename, new_filename)
        return tmp_path, new_filename
    
    # ---------- Excel -> MD (LlamaParse with fallback) ----------
    def _convert_excel_to_md(self, source_path: Path, filename: str) -> Tuple[Path, str]:
        # 1) Try LlamaParse first (if configured)
        if self._llama is not None:
            try:
                return self._llama(source_path, filename)
            except PermissionError as e:
                # auth problem: no point retrying here; fall back
                logger.error("LlamaParse auth error, falling back to local Excel conversion: %s", e)
                raise e
            except Exception as e:
                logger.error("LlamaParse failed, falling back to local Excel conversion: %s", e)
                raise e

        # 2) Local fallback using openpyxl + pandas
        return self._convert_excel_to_md_local(source_path, filename)

    def _convert_excel_to_md_local(self, source_path: Path, filename: str) -> Tuple[Path, str]:

        def get_cell_value(ws, row: int, col: int):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                return str(cell.value)
            # Handle merged cells
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return str(top_left.value) if top_left.value is not None else ""
            return ""

        with open(source_path, "rb") as f:
            wb = openpyxl.load_workbook(f, data_only=True)
            markdown_parts = []

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                max_row = ws.max_row
                max_col = ws.max_column

                data = []
                for r in range(1, max_row + 1):
                    row = []
                    for c in range(1, max_col + 1):
                        row.append(get_cell_value(ws, r, c))
                    data.append(row)

                if not data:
                    markdown_parts.append(f"### Sheet: `{sheet}`\n\n_Empty sheet._\n")
                    continue

                # Trim empty rows/cols
                df = pd.DataFrame(data[1:], columns=data[0]) if len(data) > 1 else pd.DataFrame(data)
                table_md = df.to_markdown(index=False)
                markdown_parts.append(f"### Sheet: `{sheet}`\n\n{table_md}\n")

            md_content = "\n---\n".join(markdown_parts)

        new_filename = f"{os.path.splitext(filename)[0]}.md"
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".md")
        tmp_path = Path(tmp_file.name)

        with tmp_file:
            tmp_file.write(md_content.encode("utf-8"))

        logger.debug("Excel converted to Markdown: %s -> %s", filename, new_filename)
        return tmp_path, new_filename
